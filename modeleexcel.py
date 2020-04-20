from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

tab_l_score = [5,5,5,5,5,5,5,10,15,20,25,30,35,40,45,50,55,60,65,70,75,80,85,90,95,100,110,115,120,125,130,135,140,145,150,160,165,170,175,180,185,190,195,200,210,215,220,230,240,245,250,255,260,270,275,280,290,295,300,310,315,320,325,330,340,345,350,360,365,370,380,385,390,395,400,405,410,420,425,430,440,445,450,460,465,470,475,480,485,490,495,495,495,495,495,495,495,495,495,495,495]
tab_r_score = [5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,10,15,20,25,30,35,40,45,50,60,65,70,80,85,90,95,100,110,115,120,125,130,140,145,150,160,165,170,175,180,190,195,200,210,215,220,225,230,235,240,250,255,260,265,270,280,285,290,300,305,310,320,325,330,335,340,350,355,360,365,370,380,385,390,395,400,405,410,415,420,425,430,435,445,450,455,465,470,480,485,490,495,495,495,495]

def score_toeic_l(x):
    return tab_l_score[x]

def score_toeic_r(x):
    return tab_r_score[x]


def score_toeic_l_old(x):
    if x == 0:
        return 0
    if x >= 1 and x < 17:
        return 5
    if x >= 17 and x < 28:
        return (x-15)*5
    if x == 28:
        return 70
    if x >= 29 and x < 35:
        return (x-13)*5
    if x >= 35 and x < 38:
        return (x-35)*10+115
    if x >= 38 and x < 46:
        return (x-38)*10+140
    if x == 46:
        return 205
    if x >= 47 and x < 53:
        return (x-47)*5+215
    if x >= 53 and x < 56:
        return (x-53)*5+255
    if x == 56:
        return 275
    if x >= 57 and x < 61:
        return (x-57)*5+285
    if x == 61:
        return 310
    if x >= 62 and x < 74:
        return (x-62)*5+320
    if x == 74:
        return 385
    if x >= 75 and x < 78:
        return (x-75)*5+395
    if x >= 78 and x < 89:
        return (x-78)*5+415
    if x >= 89 and x < 93:
        return (x-89)*5+475
    if x >= 93 and x < 101:
        return 495
    return -1

def score_toeic_r_old(x):
    if x == 0:
        return 0
    if x >= 1 and x < 21:
        return 5
    if x >= 21 and x < 29:
        return (x-19)*5
    if x >= 29 and x < 38:
        return (x-18)*5
    if x >= 38 and x < 46:
        return (x-38)*5+105
    if x >= 46 and x < 48:
        return (x-46)*5+155
    if x >= 48 and x < 50:
        return (x-48)*5+170
    if x >= 50 and x < 53:
        return (x-50)*10+185
    if x >= 53 and x < 55:
        return (x-53)*5+210
    if x >= 55 and x < 58:
        return (x-55)*10+220
    if x >= 58 and x < 62:
        return (x-58)*5+245
    if x >= 62 and x < 68:
        return (x-62)*5+270
    if x >= 68 and x < 70:
        return (x-68)*5+295
    if x >= 70 and x < 78:
        return (x-70)*5+310
    if x >= 78 and x < 80:
        return (x-78)*5+355
    if x >= 80 and x < 82:
        return (x-80)*5+370
    if x >= 82 and x < 87:
        return (x-82)*5+385
    if x >= 87 and x < 92:
        return (x-87)*5+415
    if x >= 92 and x < 95:
        return (x-92)*5+450
    if x >= 95 and x < 98:
        return (x-95)*5+470
    if x >= 98 and x < 101:
        return (x-98)*5+485
    return -1

def init_excel(ws1,ws2,ws3):
    ws1.title = "score"
    ws1['A1'] = "Image Nom et Prénom"
    ws1['B1'] = "Nom et Prénom"
    ws1['C1'] = "Nombre Réponses Listening"
    ws1['D1'] = "Nombre Réponses Reading"
    ws1['E1'] = "Score Listening"
    ws1['F1'] = "Score Reading"
    ws1.column_dimensions['A'].width = 120
    for i in range(0,100):
        if i < 26:
            rep = chr(ord('A') + i)
        else:
            if i < 26*2:
                rep = 'A'+chr(ord('A') + i-26)
            else:
                if i < 26*3:
                    rep = 'B'+chr(ord('A') + i-26*2)
                else:
                    rep = 'C'+chr(ord('A') + i-26*3)
        ws2[rep+str(1)] = 'Q'+str(i+1)
        ws3[rep+str(1)] = 'Q'+str(i+1)

def inserer_reponses(ws2,ws3, listerep, i):
    for j in range(0,100):
        if j < 26:
            rep = chr(ord('A') + j)
        else:
            if j < 26*2:
                rep = 'A'+chr(ord('A') + j-26)
            else:
                if j < 26*3:
                    rep = 'B'+chr(ord('A') + j-26*2)
                else:
                    rep = 'C'+chr(ord('A') + j-26*3)
        ws2[rep+str(i+2)] = chr(ord('A')+listerep[i][j]-1)
        ws3[rep+str(i+2)] = chr(ord('A')+listerep[i][j+100]-1)

def remplirerr(ws4,listeerr):
    ws4['A1'] = "feuille erronnée"
    ws4['B1'] = "erreur"
    for i in range(0,len(listeerr)):
        ws4['A'+str(i+2)] = listeerr[i][0]
        ws4['B'+str(i+2)] = listeerr[i][1]

def excel_final(nomfichier, listenom, listerep, listescore, listeerr):
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet(title="listening")
    ws3 = wb.create_sheet(title="reading")
    ws4 = wb.create_sheet(title="erreurs")
    init_excel(ws1,ws2,ws3)
    remplirerr(ws4,listeerr)
    for i in range(0,len(listenom)):
        ws1['C'+str(i+2)] = listescore[i][0]
        ws1['D'+str(i+2)] = listescore[i][1]
        ws1['E'+str(i+2)] = score_toeic_l(listescore[i][0])
        ws1['F'+str(i+2)] = score_toeic_r(listescore[i][1])
        ws1.row_dimensions[i+2].height = 140
        img = Image(listenom[i])
        ws1.add_image(img, 'A'+str(i+2))
        inserer_reponses(ws2, ws3, listerep, i)
    wb.save(filename=nomfichier)
