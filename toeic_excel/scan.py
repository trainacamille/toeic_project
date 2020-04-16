import os
import cv2
from pdf2image import convert_from_path
import numpy as np
import aide
import json


from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

def score_toeic(x):
    if x == 0:
        return 0
    if x >= 1 and x < 17:
        return 5
    if x >= 17 and x < 28:
        return (x-15)*5
    if x == 28:
        return 70
    if x >= 29 and x < 35:
        return (x-13)*5
    if x >= 36 and x < 88:
        return (x-36)*10+115
    if x >= 38 and x < 46:
        return (x-38)*10+140
    if x == 46:
        return 205
    if x >= 47 and x < 53:
        return (x-47)*5+215
    if x >= 53 and x < 56:
        return (x-53)*5+255
    if x >= 57 and x < 61:
        return (x-57)*5+285
    if x == 61:
        return 310
    if x >= 62 and x < 74:
        return (x-62)*5+320
    if x == 74:
        return 385
    if x >= 75 and x < 78:
        return (x-75)*5+395
    if x >= 78 and x < 89:
        return (x-78)*5+415
    if x >= 89 and x < 93:
        return (x-89)*5+475
    if x >= 93 and x < 101:
        return 495

def remplir_excel(nomfichier, nb_etudiant, nbr, nbl, imgnom):
    #wb = load_workbook(filename=nomfichier)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "score"
    ws2 = wb.create_sheet(title="questions")
    ws1['A1'] = "Nom et Prénom"
    ws1['B1'] = "Nombre Réponses Reading"
    ws1['C1'] = "Nombre Réponses Listening"
    ws1['D1'] = "Score Reading"
    ws1['E1'] = "Score Listening"
    ws1.column_dimensions['A'].width = 180
    ws2['A1'] = "Numéro ligne étudiant"

    #ws1 = wb.active
    img = Image(imgnom)
    ws1.add_image(img, 'A'+str(nb_etudiant+1))
    ws1['B'+str(nb_etudiant+1)] = nbr
    ws1['C'+str(nb_etudiant+1)] = nbl
    ws1['D'+str(nb_etudiant+1)] = score_toeic(nbr)
    ws1['E'+str(nb_etudiant+1)] = score_toeic(nbl)
    wb.save(filename=nomfichier)


def pdfimg(nom,dossier):
    feuille_reponse=convert_from_path(nom,dpi=200)
    for i,mat in enumerate(feuille_reponse):
        mat.save(os.path.join(dossier, 'epreuve'+str(i)+'.jpg'), 'JPEG')
    return i

def rotation(img,angle):
    rows, cols = img.shape[0], img.shape[1]
    rot=cv2.getRotationMatrix2D((cols / 2, rows / 2),angle,1)

    (cX, cY) = (cols // 2, rows // 2)
    cos = np.abs(rot[0, 0])
    sin = np.abs(rot[0, 1])
    # compute the new bounding dimensions of the image
    nW = int((rows * sin) + (cols * cos))
    nH = int((rows * cos) + (cols * sin))
    # adjust the rotation matrix to take into account translation
    rot[0, 2] += (nW / 2) - cX
    rot[1, 2] += (nH / 2) - cY

    return cv2.warpAffine(img,rot,(nW,nH))

def detectbord(img):
    #une image cany ou tresh
    contours, hierarchy = cv2.findContours(img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    ctri=sorted(contours, key=cv2.contourArea, reverse=True)
    carre=[]
    pliste=[]
    for cnt in ctri:
        ln=cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, 0.02 * ln, True)
        if(len(approx)==4):
            (x, y, w, h) = cv2.boundingRect(cnt)
            #si ca ressemble a un carré et que cest moyen (eviter les sections)
            if(w in range(20,50) and h in range(20,50)):
                #cv2.drawContours(image, [cnt], -1, (255,0,0), 2)
                carre.append(cnt)
                (tl, tr, br, bl)=aide.order_points(approx.reshape(4,2))   
                pliste.append((tr[0],tr[1]))
    #maintenant on veut retourner les coordonnées de ces points pour trouver l'ordre de rotation
    #ains que la liste des contours qui nous servira pour le zoom sur les bords
    return pliste,carre

def extraction_nom(img,i,contours,mon_dossier='correction'):
    #une image couleur    
    for cnt in contours:
        ln=cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, 0.02 * ln, True)
        if(len(approx)==4):
            (x, y, w, h) = cv2.boundingRect(cnt)
            #essayer de detecter le nom
            if(w in range(720,791) and h in range(160,201)):
                approx=approx.reshape(4,2)
                approx=aide.order_points(approx)
                nom=img[int(approx[0][1]):int(approx[2][1]),int(approx[0][0]):int(approx[2][0])]
                cv2.imwrite(mon_dossier+'/nom'+str(i)+'.jpg',nom)
                #Seul un contour aurait cette taille.
                return mon_dossier+'/nom'+str(i)+'.jpg'
    return None

def detection_sections(img,contours):
    sliste=[]
    for cnt in contours:
        ln=cv2.arcLength(cnt, True)
        approx = cv2.approxPolyDP(cnt, 0.02 * ln, True)
        if(len(approx)==4):
            (x, y, w, h) = cv2.boundingRect(cnt)
            #detecter les sections
            if(w in range(975,1150) and h in range(975,1150)):
                approx=approx.reshape(4,2)
                approx=aide.order_points(approx)
                section=img[int(approx[0][1]):int(approx[2][1]),int(approx[0][0]):int(approx[2][0])]
                sliste.append(section)
    #listening puis reading            
    return sliste

def moyenne(liste):
    somme=0
    max=-1
    for i in range(0,len(liste)):
        somme+=liste[i]
    return somme/len(liste)

def bonne_reponse(liste):
    nb=0
    ln=[]
    mean=moyenne(liste)*1.15
    for u in range(0,len(liste)):
        if(liste[u]>mean ):
            rp = u+1
            nb+=1
            ln.append(liste[u])
    if nb==1:
        return rp
    elif nb>1:
        #augmenter le seuil pour les doubles reponses
        nb=0
        mean=moyenne(liste)*1.25
        for u in range(0,len(ln)):
            if(ln[u]>mean ):
                rp = u+1
                nb+=1
        if nb==1:
            return rp
    return -1


def correction_fine(image,n):
    #prend l'image en imseuil et décale pour sauter les num questions et les bords du haut
    largeur=image.shape[1]
    hauteur=image.shape[0]-4
    #n = 5 pour le listening et 4 pour le reading(a cause de la longueur des questions)
    depart_x=int(largeur/n)+1
    img=image[7:hauteur,depart_x:largeur]
    #la diviser en 4
    nv_h=img.shape[0]
    nv_l=img.shape[1]
    proposition=[]
    pas=int(nv_l/4)
    bloc=0
    for i in range(0,4):
        prop=img[0:nv_h,bloc:bloc+pas]
        bloc+=pas
        total = cv2.countNonZero(prop)
        proposition.append(total)
    return bonne_reponse(proposition)

def recuperer_reponse(chemin):
    with open(chemin) as json_data:
        reponses = json.load(json_data)
        listening_r=reponses["Listening"][0]
        reading_r=reponses["Reading"][0]
    return listening_r,reading_r

def division_image(partie):
    hauteur=partie.shape[0]
    largeur=partie.shape[1]
    pas=int(largeur/4)-5
    questionsl=[]
    bloc=0
    for i in range (0,4):
        questionsl.append(partie[0:hauteur,bloc:bloc+pas])
        bloc+=pas

    blocl=[]
    pas=int(hauteur/25)-3
    for i in range(0,len(questionsl)):    
        depart=int(hauteur/16)+5
        hb=questionsl[i].shape[0]
        lb=questionsl[i].shape[1]
        for j in range(0,25):
            blocl.append(questionsl[i][depart:depart+pas,0:lb])
            depart=depart+pas
    choix=[]
    seuillage=[]
    if (len(blocl)!=0 ):
        for i in range(0,len(blocl)):
            possibilite=[]
            question=blocl[i]
            re,imseuil=cv2.threshold(question,0,255,cv2.THRESH_BINARY_INV+cv2.THRESH_OTSU)
            seuillage.append(imseuil)
            contours, hierarchy = cv2.findContours(imseuil, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            for c in contours:
                (x, y, w, h) = cv2.boundingRect(c)
                if (w >= 24 and h >= 24):
                    possibilite.append(c)       
            choix.append(possibilite)
    return choix,seuillage  

def detection_rep(choix,seuillage,part_r,n):
    score=0
    #1- on descend dans chacune des questions (choix)
    for e,quest in enumerate(choix):
        rponses=[]
        if(len(quest)>2):
            quest=aide.sort_contours(quest)[0]
        #Gestion erreur pas pu avoir toutes les questions exactement
        if(len(quest)!=4):
            reponse=correction_fine(seuillage[e],n)
            #print(e+1, reponse)
        #2-pour chaque question on boucle sur les 4 possibilités
        else:
            for j,c in enumerate(quest):
                mask=np.zeros(seuillage[e].shape,dtype=np.uint8)
                cv2.drawContours(mask, [c], -1, 255, -1)
                #Pour chaque possibilité compter la valeur
                mask = cv2.bitwise_and(seuillage[e], seuillage[e], mask=mask)
                total = cv2.countNonZero(mask)

                #On met a jour le total
                rponses.append(total)
            reponse=bonne_reponse(rponses)
            if(bonne_reponse(rponses)==-1):
                print(e+1,rponses)
        #Pour finir on Corrige
        #print("Reponse question "+ str(e+1)+": "+str(reponse) +" bonne reponse: "+str(part_r[str(e)]) ) 
        if(n==5):
            if reponse==part_r[str(e)]:
                score+=1
        elif n==4:
            if reponse==part_r[str(e+100)]:
                score+=1
    return score

def corriger(nom,path):
    nom='JSON/toeic_maison.json'
    #creer un repertoire dans lequel seront enregistrés les documents nécessaires à la correction
    mon_dossier='correction'
    os.mkdir(mon_dossier)
    #prendre le nom du fichier par la prof(ici statique) et le chemin créé 
    nb_pages=scan.pdfimg(path,mon_dossier)

    #Commencer le traitement pour chacune des pages recensées
    for i in range(0,nb_pages+1):
        img=cv2.imread(mon_dossier+'/epreuve'+str(i)+'.jpg',cv2.IMREAD_COLOR)

        #Si on n'a pas eu l'image
        if(not img.data):
            print('Oups, ton image n\'existe pas dis-donc!')
            continue
        #Detecter les contours pour la rotation
        #Pour cela appliquer le prétraitement de l'image et passer l'image traitée

        imgray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        contraste=cv2.convertScaleAbs(imgray,None,0.8,0)
        imflou=cv2.GaussianBlur(contraste,(5,5),0)
        imcanny=cv2.Canny(imflou,75,150,(3,3))

        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5,5))
        dest=cv2.dilate(imcanny,kernel)
        coord_bords,cont_bords=scan.detectbord(dest)

        if(len(coord_bords) != 5):
            print("Oups, je n'ai pas trouvé tous les points pour faire la bonne rotation")
            continue
        #On a trouve nos 5 carrés du bord et on va chercher la bonne rotation

        #Penser à l'écart pour une efficacité
        angle_rot=aide.rotationimage(coord_bords,80)
        imrote=scan.rotation(dest,angle_rot)
        imflou=scan.rotation(imflou,angle_rot)
        couleur=scan.rotation(img,angle_rot)

        #Apres rotation, on redétecte les bords sur l'image pour la remetre droite
        #si elle était de travers

        #(peut etre cette partie est-elle inutile mais on verra)

        coord_bords,cont_bords=scan.detectbord(imrote)
        cont=aide.sort_contours(cont_bords)[0]
        pliste=[]
        for u,cnt in enumerate(cont):
            #on a etudie la feuille et on sait que le repere au milieu est le 3eme point
            #on veut le sauter pour calibrer sur les bords
            if (u!=2):
                ln=cv2.arcLength(cnt, True)
                approx = cv2.approxPolyDP(cnt, 0.02 * ln, True)
                (tl, tr, br, bl)=aide.order_points(approx.reshape(4,2))   
                pliste.append((tr[0],tr[1]))  
        rec=aide.order_points(np.array(pliste))
        paper=aide.four_point_transform(imflou,rec)
        pdest=aide.four_point_transform(imrote,rec)
        papier=aide.four_point_transform(couleur,rec)

        #On s'attaque au nom et au listening/reading
        contours, hierarchy = cv2.findContours(pdest, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        ctri=sorted(contours, key=cv2.contourArea, reverse=False)
        chemin=scan.extraction_nom(papier,i,ctri)
        parties=scan.detection_sections(paper,ctri)

        if(len(parties)!=2):
            print(len(parties))
            print("Le listening et le reading n'ont pas étés trouves")
            continue
        listening=parties[0]
        reading=parties[1]
        listening_r,reading_r=scan.recuperer_reponse(nom)
        choix1,seuillage1=scan.division_image(listening)
        choix2,seuillage2=scan.division_image(reading)

        score_l=scan.detection_rep(choix1,seuillage1,listening_r,5)
        score_r=scan.detection_rep(choix2,seuillage2,reading_r,4)

        #ici envoyer le resultat à l'excel
        scan.remplir_excel("toeic_test.xlsx",i+1,score_r,score_l,chemin)

    print()
    print("Correction Terminee")    

    os.system("rmdir name "+mon_dossier+" /S /Q")
