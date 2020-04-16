from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

def creer_excel(nom):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "score"
    ws2 = wb.create_sheet(title="questions")
    ws1['A1'] = "Nom et Prénom"
    ws1['B1'] = "Nombre Réponses Reading"
    ws1['C1'] = "Nombre Réponses Listening"
    ws1['D1'] = "Score Reading"
    ws1['E1'] = "Score Listening"
    ws1.column_dimensions['A'].width = 100
    ws2['A1'] = "Numéro ligne étudiant"
    wb.save(filename=nom)

def score_toeic_l(x):
    if x == 0:
        return 0
    if x >= 1 and x < 17:
        return 5
    if x >= 17 and x < 28:
        return (x-15)*5
    if x == 28:
        return 70
    if x >= 29 and x < 35:
        return (x-13)*5
    if x >= 35 and x < 38:
        return (x-35)*10+115
    if x >= 38 and x < 46:
        return (x-38)*10+140
    if x == 46:
        return 205
    if x >= 47 and x < 53:
        return (x-47)*5+215
    if x >= 53 and x < 56:
        return (x-53)*5+255
    if x == 56:
        return 275
    if x >= 57 and x < 61:
        return (x-57)*5+285
    if x == 61:
        return 310
    if x >= 62 and x < 74:
        return (x-62)*5+320
    if x == 74:
        return 385
    if x >= 75 and x < 78:
        return (x-75)*5+395
    if x >= 78 and x < 89:
        return (x-78)*5+415
    if x >= 89 and x < 93:
        return (x-89)*5+475
    if x >= 93 and x < 101:
        return 495
    return -1

def score_toeic_r(x):
    if x == 0:
        return 0
    if x >= 1 and x < 21:
        return 5
    if x >= 21 and x < 29:
        return (x-19)*5
    if x >= 29 and x < 38:
        return (x-18)*5
    if x >= 38 and x < 46:
        return (x-38)*5+105
    if x >= 46 and x < 48:
        return (x-46)*5+155
    if x >= 48 and x < 50:
        return (x-48)*5+170
    if x >= 50 and x < 53:
        return (x-50)*10+185
    if x >= 53 and x < 55:
        return (x-53)*5+210
    if x >= 55 and x < 58:
        return (x-55)*10+220
    if x >= 58 and x < 62:
        return (x-58)*5+245
    if x >= 62 and x < 68:
        return (x-62)*5+270
    if x >= 68 and x < 70:
        return (x-68)*5+295
    if x >= 70 and x < 78:
        return (x-70)*5+310
    if x >= 78 and x < 80:
        return (x-78)*5+355
    if x >= 80 and x < 82:
        return (x-80)*5+370
    if x >= 82 and x < 87:
        return (x-82)*5+385
    if x >= 87 and x < 92:
        return (x-87)*5+415
    if x >= 92 and x < 95:
        return (x-92)*5+450
    if x >= 95 and x < 98:
        return (x-95)*5+470
    if x >= 98 and x < 101:
        return (x-98)*5+485
    return -1

def remplir_excel(nomfichier, nb_etudiant, nbr, nbl, imgnom):
    #wb = load_workbook(filename=nomfichier)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "score"
    ws2 = wb.create_sheet(title="questions")
    ws1['A1'] = "Nom et Prénom"
    ws1['B1'] = "Nombre Réponses Reading"
    ws1['C1'] = "Nombre Réponses Listening"
    ws1['D1'] = "Score Reading"
    ws1['E1'] = "Score Listening"
    ws1.column_dimensions['A'].width = 180
    ws1.row_dimensions[nb_etudiant+1].height = 240
    ws2['A1'] = "Numéro ligne étudiant"

    #ws1 = wb.active
    img = Image(imgnom)
    ws1.add_image(img, 'A'+str(nb_etudiant+1))
    ws1['B'+str(nb_etudiant+1)] = nbr
    ws1['C'+str(nb_etudiant+1)] = nbl
    ws1['D'+str(nb_etudiant+1)] = score_toeic_r(nbr)
    ws1['E'+str(nb_etudiant+1)] = score_toeic_l(nbl)
    wb.save(filename=nomfichier)

def init_excel(ws1,ws2,ws3):
    ws1.title = "score"
    ws1['A1'] = "Image Nom et Prénom"
    ws1['B1'] = "Nom et Prénom"
    ws1['C1'] = "Nombre Réponses Reading"
    ws1['D1'] = "Nombre Réponses Listening"
    ws1['E1'] = "Score Reading"
    ws1['F1'] = "Score Listening"
    ws1.column_dimensions['A'].width = 180
    for i in range(0,100):
        if i < 26:
            rep = chr(ord('A') + i)
        else:
            if i < 26*2:
                rep = 'A'+chr(ord('A') + i-26)
            else:
                if i < 26*3:
                    rep = 'B'+chr(ord('A') + i-26*2)
                else:
                    rep = 'C'+chr(ord('A') + i-26*3)
        ws2[rep+str(1)] = 'Q'+str(i+1)
        ws3[rep+str(1)] = 'Q'+str(i+1)

def inserer_reponses(ws2,ws3, listerep, i):
    for j in range(0,100):
        if j < 26:
            rep = chr(ord('A') + j)
        else:
            if j < 26*2:
                rep = 'A'+chr(ord('A') + j-26)
            else:
                if j < 26*3:
                    rep = 'B'+chr(ord('A') + j-26*2)
                else:
                    rep = 'C'+chr(ord('A') + j-26*3)
        ws2[rep+str(i+2)] = 'A'
        ws3[rep+str(i+2)] = 'B'
        '''ws2[rep+str(j+1)] = listerep[i][j]
        ws3[rep+str(j+1)] = listerep[i][j+100]'''

def excel_final(nomfichier, listenom, listerep, listescore):
    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet(title="listening")
    ws3 = wb.create_sheet(title="reading")
    init_excel(ws1,ws2,ws3)
    
    for i in range(0,len(listenom)):
        ws1['C'+str(i+2)] = listescore[i][0]
        ws1['D'+str(i+2)] = listescore[i][1]
        ws1['E'+str(i+2)] = score_toeic_r(listescore[i][0])
        ws1['F'+str(i+2)] = score_toeic_l(listescore[i][1])
        ws1.row_dimensions[i+2].height = 240
        img = Image(listenom[i])
        ws1.add_image(img, 'A'+str(i+2))
        inserer_reponses(ws2, ws3, listerep, i)
    wb.save(filename=nomfichier)


l = ["toeic_excel/nom.jpg","toeic_excel/nom.jpg"]
l2 = [['A','A'],['B','B']]
l3 = [[100,50],[10,20]]

excel_final("toeic_excel/text2.xlsx",l,l2,l3)

'''
for i in range(0,101):
    print (i,score_toeic_l(i), score_toeic_r(i))
'''